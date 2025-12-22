import pandas as pd
import numpy as np
from scipy import stats

"""
Calculate Statistical Significance of Correlations for Vietnamese Enterprise Survey

This script:
1. Loads pivoted data with 23 enterprise segments
2. Calculates t-statistics and p-values for key correlations
3. Determines significance levels (***/**/*/ ns)
4. Classifies correlation strength (Very strong/Strong/Moderate/Weak)
5. Provides interpretations combining strength and significance
6. Exports formatted Excel tables

Interpretation format:
"[Strength] [direction] correlation, [significance]"

Examples:
- "Very strong positive correlation, highly significant"
- "Strong negative correlation, significant"
- "Moderate positive correlation, significant"
- "Weak negative correlation, not significant"

Output files:
- Table4_Correlation_Significance.xlsx: Main table with interpretations and notes
- Table4_For_Paper.xlsx: Compact version for publication (no interpretation column)
- All_Correlations_Detailed.xlsx: Complete analysis of all correlation pairs
"""

# Load the pivoted data (the one with 23 rows and 10 indicator columns)
# If you already have it as CSV, load it directly
# Otherwise, follow the pivoting process from the main analysis script

# For demonstration, Show both methods:

# METHOD 1: If we have the pivoted CSV already
try:
    pivoted = pd.read_csv('Pivoted_Data.csv')
    print("Loaded pivoted data from CSV")
except FileNotFoundError:
    print("Pivoted_Data.csv not found. Creating from raw data...")
    
    # METHOD 2: Create from raw data
    data = pd.read_csv('vietnam_data_clean.csv')
    
    # Replace '.' with NaN and convert value to numeric
    data['value'] = data['value'].replace('.', np.nan)
    data['value'] = pd.to_numeric(data['value'], errors='coerce')
    
    # Indicators for correlation
    corr_indicators = ['t5', 't7', 't9', 'perf1', 'perf2', 'perf3', 
                       'bready_t1', 'fin14', 'bready_fin28', 'bready_fin31']
    
    # Collect all unique subcuts
    all_groups = data[data['subcut'] != '']
    
    # Select and drop duplicates
    filtered = all_groups[all_groups['indicator'].isin(corr_indicators)].drop_duplicates(
        subset=['cut', 'subcut', 'indicator'], keep='first')
    
    # Pivot: subgroups as index, indicators as columns
    pivoted = filtered.pivot(index=['cut', 'subcut'], columns='indicator', values='value').reset_index()
    
    # Drop rows with all NaN in indicators
    pivoted = pivoted.dropna(how='all', subset=corr_indicators)
    
    # Save for future use
    pivoted.to_csv('Pivoted_Data.csv', index=False)
    print("Created and saved Pivoted_Data.csv")

# Define the 10 indicators
corr_indicators = ['t5', 't7', 't9', 'perf1', 'perf2', 'perf3', 
                   'bready_t1', 'fin14', 'bready_fin28', 'bready_fin31']

# Indicator descriptions for better table readability
indicator_names = {
    't5': 'Website adoption',
    't7': 'Product innovation',
    't9': 'Process innovation',
    'perf1': 'Sales growth',
    'perf2': 'Employment growth',
    'perf3': 'Productivity growth',
    'bready_t1': 'Quality certification',
    'fin14': 'Bank loan access',
    'bready_fin28': 'E-payment sales',
    'bready_fin31': 'E-payment purchases'
}

# Calculate correlation matrix
corr_matrix = pivoted[corr_indicators].corr()

print(f"\nNumber of segments (observations): {len(pivoted)}")
print(f"Number of indicators: {len(corr_indicators)}")

# Function to calculate t-statistic for correlation
def calculate_t_statistic(r, n):
    """
    Calculate t-statistic for Pearson correlation
    
    Parameters:
    r: correlation coefficient
    n: sample size (number of observations)
    
    Returns:
    t: t-statistic
    """
    print("Running calculate_t_statistic with r =", r, "and n =", n)
    
    if abs(r) == 1:
        return np.inf if r > 0 else -np.inf
    
    t = r * np.sqrt((n - 2) / (1 - r**2))
    return t

# Function to get p-value from t-statistic
def get_p_value(t, df):
    """
    Calculate two-tailed p-value from t-statistic
    
    Parameters:
    t: t-statistic
    df: degrees of freedom
    
    Returns:
    p: p-value (two-tailed)
    """
    if np.isinf(t):
        return 0.0
    
    # Two-tailed test
    p = 2 * (1 - stats.t.cdf(abs(t), df))
    return p

# Function to get significance symbol
def get_significance(p):
    """
    Return significance symbol based on p-value
    """
    if p < 0.001:
        return '***'
    elif p < 0.01:
        return '**'
    elif p < 0.05:
        return '*'
    else:
        return 'ns'

# Function to get correlation strength description
def get_strength_description(r, p):
    """
    Return a complete description of correlation strength and significance
    
    Strength categories based on |r|:
    - Very strong: |r| > 0.70
    - Strong: 0.50 < |r| ≤ 0.70
    - Moderate: 0.30 < |r| ≤ 0.50
    - Weak: |r| ≤ 0.30
    
    Significance categories:
    - Highly significant: p < 0.001
    - Significant: p < 0.01
    - Marginally significant: p < 0.05
    - Not significant: p ≥ 0.05
    """
    abs_r = abs(r)
    
    # Determine strength
    if abs_r > 0.70:
        strength = "Very strong"
    elif abs_r > 0.50:
        strength = "Strong"
    elif abs_r > 0.30:
        strength = "Moderate"
    else:
        strength = "Weak"
    
    # Determine direction
    if r > 0:
        direction = "positive"
    else:
        direction = "negative"
    
    # Determine significance
    if p < 0.001:
        significance = "highly significant"
    elif p < 0.01:
        significance = "significant"
    elif p < 0.05:
        significance = "marginally significant"
    else:
        significance = "not significant"
    
    # Combine into description
    return f"{strength} {direction} correlation, {significance}"

# Number of observations (segments) and degrees of freedom
n = len(pivoted)
df = n - 2

print(f"Degrees of freedom: {df}")

# Critical value at 5% significance level (two-tailed)
critical_t = stats.t.ppf(0.975, df)  # 0.975 for two-tailed at 0.05
print(f"Critical t-value (α=0.05, two-tailed): {critical_t:.2f}\n")

# Create list to store results
results = []

# Define key relationships to analyze (based on your Table 4)
key_relationships = [
    ('bready_fin28', 'perf3', 'E-payment sales ↔ Productivity growth'),
    ('bready_fin31', 'perf3', 'E-payment purchases ↔ Productivity growth'),
    ('perf1', 'perf3', 'Sales growth ↔ Productivity growth'),
    ('bready_fin28', 'bready_fin31', 'E-payment sales ↔ E-payment purchases'),
    ('bready_fin28', 'perf1', 'E-payment sales ↔ Sales growth'),
    ('t7', 't9', 'Product innovation ↔ Process innovation'),
    ('t9', 'fin14', 'Process innovation ↔ Bank loan access'),
    ('t5', 'bready_t1', 'Website ↔ Quality certification'),
    ('bready_fin31', 'perf1', 'E-payment purchases ↔ Sales growth'),
    ('t5', 'perf3', 'Website ↔ Productivity growth'),
    ('t5', 'bready_fin28', 'Website ↔ E-payment sales'),
    ('t5', 'perf2', 'Website ↔ Employment growth'),
    ('perf2', 'perf3', 'Employment growth ↔ Productivity growth'),
    ('bready_fin31', 'perf2', 'E-payment purchases ↔ Employment growth'),
    ('t7', 'perf3', 'Product innovation ↔ Productivity growth'),
    ('t7', 'perf1', 'Product innovation ↔ Sales growth'),
    ('t5', 't7', 'Website ↔ Product innovation'),
]

# Calculate statistics for each relationship
for ind1, ind2, description in key_relationships:
    # r = corr_matrix.loc[ind1, ind2]
    # t = calculate_t_statistic(r, n)
    # p = get_p_value(t, df)
    # sig = get_significance(p)
    
    # Get strength and significance description
    r = corr_matrix.loc[ind1, ind2]
    r = round(r, 2) # Round r for display
    t = calculate_t_statistic(r, n)
    p = get_p_value(t, df)
    sig = get_significance(p)
    interpretation = get_strength_description(r, p)
    
    # Format p-value for display
    if p < 0.001:
        p_display = '<0.001'
    elif p < 0.01:
        p_display = '<0.01'
    elif p < 0.05:
        p_display = '<0.05'
    else:
        p_display = f'{p:.3f}'
    
    results.append({
        'Relationship': description,
        'r': round(r, 2),
        't-value': round(t, 2),
        'p-value': p_display,
        'p_numeric': p,  # Keep numeric for sorting
        'Significance': sig,
        'Interpretation': interpretation
    })

# Create DataFrame
results_df = pd.DataFrame(results)

# Sort by absolute t-value (strongest correlations first)
results_df['abs_t'] = results_df['t-value'].abs()
results_df = results_df.sort_values('abs_t', ascending=False)
results_df = results_df.drop('abs_t', axis=1)

# Remove p_numeric column for final display
results_df_display = results_df.drop('p_numeric', axis=1)

# Set pandas display options for better table formatting
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', 80)

# Print the table
print("="*150)
print("TABLE 4. STATISTICAL SIGNIFICANCE OF KEY CORRELATIONS")
print("="*150)
print(results_df_display.to_string(index=False))
print("\n" + "="*150)
print(f"Note: *** p<0.001 (highly significant), ** p<0.01 (significant),")
print(f"      * p<0.05 (marginally significant), ns = not significant")
print(f"Critical value: |t| > {critical_t:.2f}")
print(f"Based on {n} enterprise segments with {df} degrees of freedom")
print("")
print("Correlation Strength Categories:")
print("  - Very strong: |r| > 0.70")
print("  - Strong: 0.50 < |r| ≤ 0.70")
print("  - Moderate: 0.30 < |r| ≤ 0.50")
print("  - Weak: |r| ≤ 0.30")
print("="*150)

# Export to Excel with formatting
with pd.ExcelWriter('Table4_Correlation_Significance.xlsx', engine='openpyxl') as writer:
    results_df_display.to_excel(writer, sheet_name='Significance Table', index=False)
    
    # Get the worksheet to adjust column widths
    worksheet = writer.sheets['Significance Table']
    
    # Adjust column widths for better readability
    worksheet.column_dimensions['A'].width = 45  # Relationship
    worksheet.column_dimensions['B'].width = 8   # r
    worksheet.column_dimensions['C'].width = 10  # t-value
    worksheet.column_dimensions['D'].width = 10  # p-value
    worksheet.column_dimensions['E'].width = 15  # Significance
    worksheet.column_dimensions['F'].width = 50  # Interpretation (shorter now)
    
    # Add header formatting
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Enable text wrapping for interpretation column
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add notes below the table
    notes_row = worksheet.max_row + 3
    worksheet[f'A{notes_row}'] = 'Notes:'
    worksheet[f'A{notes_row}'].font = Font(bold=True)
    
    worksheet[f'A{notes_row + 1}'] = 'Significance levels: *** p<0.001 (highly significant), ** p<0.01 (significant), * p<0.05 (marginally significant), ns = not significant'
    worksheet[f'A{notes_row + 2}'] = f'Critical value: |t| > {critical_t:.2f} (based on {n} segments, {df} degrees of freedom)'
    worksheet[f'A{notes_row + 3}'] = ''
    worksheet[f'A{notes_row + 4}'] = 'Correlation Strength Categories:'
    worksheet[f'A{notes_row + 4}'].font = Font(bold=True)
    worksheet[f'A{notes_row + 5}'] = '  - Very strong: |r| > 0.70'
    worksheet[f'A{notes_row + 6}'] = '  - Strong: 0.50 < |r| ≤ 0.70'
    worksheet[f'A{notes_row + 7}'] = '  - Moderate: 0.30 < |r| ≤ 0.50'
    worksheet[f'A{notes_row + 8}'] = '  - Weak: |r| ≤ 0.30'
    
    # Merge cells for notes
    for row_num in range(notes_row + 1, notes_row + 9):
        worksheet.merge_cells(f'A{row_num}:F{row_num}')
    
print("\n✓ Table exported to: Table4_Correlation_Significance.xlsx")
print("  (Includes strength categories and interpretations)")

# Create a simplified version for paper (without interpretation column)
results_paper = results_df_display.drop('Interpretation', axis=1)
with pd.ExcelWriter('Table4_For_Paper.xlsx', engine='openpyxl') as writer:
    results_paper.to_excel(writer, sheet_name='Table 4', index=False)
    
    # Get the worksheet to adjust column widths
    worksheet = writer.sheets['Table 4']
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 45  # Relationship
    worksheet.column_dimensions['B'].width = 8   # r
    worksheet.column_dimensions['C'].width = 10  # t-value
    worksheet.column_dimensions['D'].width = 10  # p-value
    worksheet.column_dimensions['E'].width = 15  # Significance
    
    # Add header formatting
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Add borders to all cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

print("✓ Paper-ready version exported to: Table4_For_Paper.xlsx")
print("  (Compact format without interpretations for publication)")

# Optional: Create a more detailed analysis of ALL correlations
print("\n" + "="*150)
print("DETAILED CORRELATION MATRIX WITH SIGNIFICANCE")
print("="*150)

detailed_results = []

for i, ind1 in enumerate(corr_indicators):
    for j, ind2 in enumerate(corr_indicators):
        if i < j:  # Only upper triangle (avoid duplicates)
            r = corr_matrix.loc[ind1, ind2]
            r = round(r, 2) # Round r for display
            t = calculate_t_statistic(r, n)
            p = get_p_value(t, df)
            sig = get_significance(p)
            
            detailed_results.append({
                'Indicator 1': ind1,
                'Name 1': indicator_names[ind1],
                'Indicator 2': ind2,
                'Name 2': indicator_names[ind2],
                'r': round(r, 2),
                't': round(t, 2),
                'p': p,
                'sig': sig,
                'significant': 'Yes' if abs(t) > critical_t else 'No'
            })

detailed_df = pd.DataFrame(detailed_results)
detailed_df = detailed_df.sort_values('t', key=abs, ascending=False)

# Export detailed analysis
detailed_df.to_excel('All_Correlations_Detailed.xlsx', index=False)
print(f"\n✓ Detailed analysis of all {len(detailed_df)} correlation pairs exported to:")
print("  All_Correlations_Detailed.xlsx")

# Summary statistics
print("\n" + "="*150)
print("SUMMARY STATISTICS")
print("="*150)
print(f"Total correlation pairs analyzed: {len(detailed_df)}")
print(f"Highly significant (p<0.001): {len(detailed_df[detailed_df['sig'] == '***'])}")
print(f"Significant (p<0.01): {len(detailed_df[detailed_df['sig'] == '**'])}")
print(f"Marginally significant (p<0.05): {len(detailed_df[detailed_df['sig'] == '*'])}")
print(f"Not significant (p≥0.05): {len(detailed_df[detailed_df['sig'] == 'ns'])}")

# Add strength distribution for key relationships
print("\nKey Relationships by Strength:")
strength_counts = {'Very strong': 0, 'Strong': 0, 'Moderate': 0, 'Weak': 0}
for _, row in results_df.iterrows():
    r_val = row['r']
    abs_r = abs(r_val)
    if abs_r > 0.70:
        strength_counts['Very strong'] += 1
    elif abs_r > 0.50:
        strength_counts['Strong'] += 1
    elif abs_r > 0.30:
        strength_counts['Moderate'] += 1
    else:
        strength_counts['Weak'] += 1

for strength, count in strength_counts.items():
    print(f"  {strength}: {count}")

# Show strongest correlations
print("\n" + "="*150)
print("TOP 10 STRONGEST CORRELATIONS")
print("="*150)
top10 = detailed_df.head(10)[['Name 1', 'Name 2', 'r', 't', 'sig']]
print(top10.to_string(index=False))

print("\n" + "="*150)
print("ANALYSIS COMPLETE!")
print("="*150)
print("\nFiles created:")
print("1. Table4_Correlation_Significance.xlsx - Full table with interpretations")
print("2. Table4_For_Paper.xlsx - Compact version for publication")
print("3. All_Correlations_Detailed.xlsx - Complete analysis of all correlation pairs")
print("="*150)